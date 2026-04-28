# -*- coding: utf-8 -*-

import json
from typing import Dict, List, Optional
import pandas as pd
from openpyxl.styles import Font
from openpyxl import load_workbook
import logging

class ExcelPack:
    # Excel列结构配置
    COLUMNS = {
        'name': 0,
        'run': 1,
        'token': 2,
        'url': 3,
        'data': 4,
        'method': 5,
        'type': 6,
        'header': 7,
        'code': 8,
        'isSuccess': 9,
        'result': 10,
    }

    # 结果样式配置
    STYLES = {
        'pass': (3, '00FF00'),  # 绿色
        'fail': (4, 'FF0000'),  # 红色
        'skip': (6, '800080'),  # 紫色
    }

    def __init__(self, file_name: str, api_host: str, token: str):
        self.file_name = file_name
        self._df = None
        self.api_host = api_host
        self.token = token

    @property
    def df(self) -> pd.DataFrame:
        if self._df is None:
            self._df = pd.read_excel(self.file_name)
            logging.info(f"成功读取Excel文件: {self.file_name}")
        return self._df

    def _get_cell_value(self, row: int, col_name: str) -> str:
        """统一的单元格获取方法"""
        value = self.df.iloc[row, self.COLUMNS[col_name]]
        return str(value).strip() if pd.notna(value) else ''

    def _write_cell(self, row: int, col_name: str, value: str, style: Optional[int] = None):
        """增量更新单元格内容,保留其他数据格式"""

        # 更新DataFrame中的值
        col = self.COLUMNS[col_name]
        self.df.iloc[row, col] = value

        # 直接加载工作簿进行更新
        workbook = load_workbook(self.file_name)
        worksheet = workbook.active

        # 写入新值
        cell = worksheet.cell(row=row + 2, column=col + 1)
        cell.value = value

        # 应用样式
        if style and isinstance(style, (int, str)):
            style_config = self.STYLES.get(str(style))
            if style_config and len(style_config) >= 2:
                color = style_config[1]
                cell.font = Font(bold=True, color=color)
            else:
                logging.warning(f"无效的样式配置: {style}")

        # 保存工作簿
        workbook.save(self.file_name)

    def load_test_cases(self) -> List[Dict]:
        """加载测试用例"""
        test_cases = []

        for row in range(1, len(self.df)):
            headers = eval(self._get_cell_value(row, 'header')) if self._get_cell_value(row, 'header') else {}
            if self._get_cell_value(row, 'token').lower() == 'yes':
                headers['token'] = self.token
            
            case = {
                'name': self._get_cell_value(row, 'name'),
                'run': self._get_cell_value(row, 'run'),
                'token': self._get_cell_value(row, 'token'),
                'type': self._get_cell_value(row, 'type'),
                'request': {
                    'url': f"{self.api_host}{self._get_cell_value(row, 'url')}",
                    'method': self._get_cell_value(row, 'method'),
                    'headers': headers,
                    'data': json.loads(self._get_cell_value(row, 'data')) if self._get_cell_value(row, 'data') else None,
                },
                'expected': {
                    'code': int(self._get_cell_value(row, 'code')),
                    'isSuccess': bool(self._get_cell_value(row, 'isSuccess')),
                },
                'result': {
                    'result': self._get_cell_value(row, 'result'),
                },
                'row': row,
            }
            test_cases.append(case)
        return test_cases

    def write_to_excel(self, row: int, result: Dict) -> None:
        """写入结果到Excel"""
        try:
            # 更新数据
            self._write_cell(row, 'result', result['result'], result['result'])
            logging.info(f"测试结果写入成功: {result['name']}")
        except Exception as e:
            logging.error(f"写入测试结果失败: {str(e)}")
            logging.exception(e)
