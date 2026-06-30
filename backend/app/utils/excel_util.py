import io
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


class ExcelUtil:
    """Excel 模板生成与列表导出（openpyxl）。"""

    @staticmethod
    def read_excel_to_dicts(contents: bytes) -> list[dict[str, Any]]:
        """读取 Excel 文件字节，返回字典列表（首行为列名）。"""
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        result: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i] is not None:
                    row_dict[headers[i]] = val
            if row_dict:
                result.append(row_dict)
        wb.close()
        return result

    @classmethod
    def __mapping_list(cls, list_data: list[dict[str, Any]], mapping_dict: dict) -> list[dict[str, Any]]:
        """
        将列表数据中的字段名映射为对应的中文字段名。

        参数:
        - list_data: 数据列表。
        - mapping_dict: 字段名映射字典 {英文key: 中文表头}。

        返回:
        - list[dict]: 映射后的数据列表 [{中文表头: value}]。
        """
        return [{mapping_dict.get(key): item.get(key) for key in mapping_dict} for item in list_data]

    @classmethod
    def get_excel_template(
        cls,
        header_list: list[str],
        selector_header_list: list[str],
        option_list: list[dict[str, list[str]]],
    ) -> bytes:
        """
        生成 Excel 模板文件。

        参数:
        - header_list: 表头列表。
        - selector_header_list: 需要设置下拉选择的表头列表。
        - option_list: 下拉选项配置列表。

        返回:
        - bytes: Excel 文件的二进制数据。
        """
        wb = Workbook()
        ws = wb.active
        if not ws:
            raise ValueError("不存在活动工作表")

        header_fill = PatternFill(start_color="ababab", end_color="ababab", fill_type="solid")

        for col_num, header in enumerate(header_list, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header  # pyright: ignore[reportAttributeAccessIssue]
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_num)].width = 12

        for selector_header in selector_header_list:
            col_idx = header_list.index(selector_header) + 1
            header_options = next(
                (opt.get(selector_header) for opt in option_list if selector_header in opt),
                [],
            )
            if header_options:
                dv = DataValidation(type="list", formula1=f'"{",".join(header_options)}"')
                dv.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}1048576")
                ws.add_data_validation(dv)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @classmethod
    def export_list2excel(cls, list_data: list[dict[str, Any]], mapping_dict: dict) -> bytes:
        """
        将列表数据导出为 Excel 文件。

        参数:
        - list_data: 要导出的数据列表。
        - mapping_dict: 字段名映射字典 {英文key: 中文表头}。

        返回:
        - bytes: Excel 文件的二进制数据。
        """
        mapping_data = cls.__mapping_list(list_data, mapping_dict)
        headers = list(mapping_dict.values())

        wb = Workbook()
        ws = wb.active
        if not ws:
            raise ValueError("不存在活动工作表")

        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, row_data in enumerate(mapping_data, 2):
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row_data.get(header))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
